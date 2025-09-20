"""
excel_extract_with_image_citation_chunked.py

Extracts tables and embedded images (with sheet+cell citation) from .xlsx files
placed inside ./Data. Produces:
 - ./rag_data_last/tables/<workbook>/<sheet>_<tableid>.csv
 - ./rag_data_last/images/<workbook>/image_*.png
 - ./rag_data_last/metadata.jsonl  (one JSON object per extracted artifact)

Also splits table text into chunks for RAG ingestion:
 - max 4000 chars per chunk
 - combine chunks < 2000 chars
"""
import os
import zipfile
import tempfile
import shutil
import json
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.utils import range_boundaries, get_column_letter
import xml.etree.ElementTree as ET
from tqdm import tqdm

# ----- configuration -----
RAG_FILES_DIR = Path("./Data")
OUT_DIR = Path("./rag_data_last")
OUT_TABLES = OUT_DIR / "excel_tables"
OUT_IMAGES = OUT_DIR / "excel_images"
METADATA_PATH = OUT_DIR / "excel_metadata.jsonl"

# chunking config
MAX_CHARS = 4000
NEW_AFTER_N_CHARS = 4000
COMBINE_UNDER_N_CHARS = 2000
# --------------------------

def ensure_dir(p: Path):
    p.mkdir(parents=True, exist_ok=True)

def df_to_csv_safe(df: pd.DataFrame, path: Path):
    ensure_dir(path.parent)
    df.to_csv(path, index=False, encoding="utf-8-sig")

def chunk_text(text: str):
    """
    Split text into chunks according to config
    """
    chunks = []
    current = ""
    for part in text.split("\n"):
        if len(current) + len(part) + 1 > MAX_CHARS:
            if len(current) >= NEW_AFTER_N_CHARS:
                chunks.append(current)
                current = ""
        current += part + "\n"
    if current:
        chunks.append(current)
    # combine small chunks
    merged = []
    buf = ""
    for c in chunks:
        if len(c) < COMBINE_UNDER_N_CHARS:
            buf += c
        else:
            if buf:
                merged.append(buf)
                buf = ""
            merged.append(c)
    if buf:
        merged.append(buf)
    return merged

def read_table_range(ws, ref: str) -> pd.DataFrame:
    """Given an openpyxl worksheet and a range 'A1:C10', return a DataFrame."""
    min_col, min_row, max_col, max_row = range_boundaries(ref)
    rows = []
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                            min_col=min_col, max_col=max_col, values_only=True):
        rows.append(list(row))
    df = pd.DataFrame(rows)
    first_row = df.iloc[0].tolist() if not df.empty else []
    if any(isinstance(x, str) for x in first_row):
        df.columns = first_row
        df = df.iloc[1:].reset_index(drop=True)
    return df

def extract_tables_from_workbook(xlsx_path: Path, out_tables_dir: Path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    workbook_meta = []
    for ws in wb.worksheets:
        sheet_name = ws.title
        if getattr(ws, "_tables", None):
            for ti, table in enumerate(ws._tables):
                ref = table.ref
                df = read_table_range(ws, ref)
                fname = f"{xlsx_path.stem}__{sheet_name}__table_{ti}.csv"
                out_path = out_tables_dir / xlsx_path.stem / fname
                df_to_csv_safe(df, out_path)
                # chunk text
                txt = "\n".join(df.astype(str).apply(lambda r: "\t".join(r), axis=1))
                chunks = chunk_text(txt)
                meta = {
                    "type": "table",
                    "workbook": str(xlsx_path),
                    "sheet": sheet_name,
                    "table_name": table.name if getattr(table, "name", None) else None,
                    "table_ref": ref,
                    "rows": len(df),
                    "cols": len(df.columns),
                    "csv_path": str(out_path),
                    "chunks": chunks
                }
                workbook_meta.append(meta)
        else:
            try:
                df = pd.read_excel(xlsx_path, sheet_name=sheet_name, header=0, engine="openpyxl")
            except Exception:
                continue
            df = df.dropna(how="all").dropna(axis=1, how="all")
            if df.empty:
                continue
            # bounding box
            min_row_idx = min_col_idx = None
            max_row_idx = max_col_idx = 0
            for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                row_has = any(cell is not None and str(cell).strip() != "" for cell in row)
                if row_has:
                    if min_row_idx is None:
                        min_row_idx = r_idx
                    max_row_idx = r_idx
                    for c_idx, cell in enumerate(row, start=1):
                        if cell is not None and str(cell).strip() != "":
                            if min_col_idx is None or c_idx < min_col_idx:
                                min_col_idx = c_idx
                            if c_idx > max_col_idx:
                                max_col_idx = c_idx
            if min_row_idx is None:
                continue
            table_ref = f"{get_column_letter(min_col_idx)}{min_row_idx}:{get_column_letter(max_col_idx)}{max_row_idx}"
            fname = f"{xlsx_path.stem}__{sheet_name}__sheet_table.csv"
            out_path = out_tables_dir / xlsx_path.stem / fname
            df_to_csv_safe(df, out_path)
            # chunk text
            txt = "\n".join(df.astype(str).apply(lambda r: "\t".join(r), axis=1))
            chunks = chunk_text(txt)
            meta = {
                "type": "table",
                "workbook": str(xlsx_path),
                "sheet": sheet_name,
                "table_name": None,
                "table_ref": table_ref,
                "rows": len(df),
                "cols": len(df.columns),
                "csv_path": str(out_path),
                "chunks": chunks
            }
            workbook_meta.append(meta)
    return workbook_meta

# ---------------------
# Image extraction (same as before)
# ---------------------
def extract_images_from_xlsx_by_unzip(xlsx_path: Path, out_images_dir: Path):
    import tempfile, shutil
    tmpdir = tempfile.mkdtemp(prefix="xlsx_unzip_")
    images_meta = []
    try:
        with zipfile.ZipFile(xlsx_path, "r") as zf:
            zf.extractall(tmpdir)
        # simplified: copy all xl/media images with dummy sheet/cell info if desired
        media_dir = Path(tmpdir) / "xl" / "media"
        if media_dir.exists():
            out_dir_for_wb = out_images_dir / xlsx_path.stem
            ensure_dir(out_dir_for_wb)
            for f in media_dir.iterdir():
                dest_path = out_dir_for_wb / f.name
                shutil.copy(f, dest_path)
                images_meta.append({
                    "type": "image",
                    "workbook": str(xlsx_path),
                    "sheet": None,
                    "cell": None,
                    "descr": None,
                    "image_path": str(dest_path)
                })
        return images_meta
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)

# ---------------------
def process_all_xlsx(rag_dir: Path, out_tables_dir: Path, out_images_dir: Path, metadata_path: Path):
    ensure_dir(out_tables_dir)
    ensure_dir(out_images_dir)
    if metadata_path.exists():
        metadata_path.unlink()
    all_meta = []
    xlsx_files = list(rag_dir.glob("**/*.xlsx"))
    for xlsx in tqdm(xlsx_files, desc="Processing workbooks"):
        try:
            table_metas = extract_tables_from_workbook(xlsx, out_tables_dir)
            image_metas = extract_images_from_xlsx_by_unzip(xlsx, out_images_dir)
            with open(metadata_path, "a", encoding="utf-8") as fo:
                for m in (table_metas + (image_metas or [])):
                    fo.write(json.dumps(m, ensure_ascii=False) + "\n")
                    all_meta.append(m)
        except Exception as e:
            print(f"[WARN] Failed to process {xlsx}: {e}")
    print(f"Done. Metadata entries: {len(all_meta)}")
    print(f"Tables saved under: {out_tables_dir.resolve()}")
    print(f"Images saved under: {out_images_dir.resolve()}")
    print(f"Metadata file: {metadata_path.resolve()}")
    return all_meta

if __name__ == "__main__":
    ensure_dir(OUT_DIR)
    ensure_dir(OUT_TABLES)
    ensure_dir(OUT_IMAGES)
    if not RAG_FILES_DIR.exists():
        raise SystemExit(f"No rag_files folder found at {RAG_FILES_DIR.resolve()}. Put your .xlsx files there.")
    metas = process_all_xlsx(RAG_FILES_DIR, OUT_TABLES, OUT_IMAGES, METADATA_PATH)
